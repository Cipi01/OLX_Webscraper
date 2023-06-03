import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv("D:/P/Webscrapers/BD/OLX_AUTO/CSV/OLX_MOTO_01_06_23.csv")

average_prices = df.groupby(['Brand', 'Model', 'AnFabr'])['Pret'].mean().round().reset_index()
average_prices = average_prices.rename(columns={'Pret': 'AvgPret'})
average_prices = average_prices[average_prices['Model'] != 'altul']

xlsx_file = "OLX_MOTO_29_05_23.xlsx"

book = load_workbook(xlsx_file)
if 'test' not in book.sheetnames:
    book.create_sheet('test')

sheet = book['test']
sheet['A1'] = "TEST TEXT TO TEST"
sheet.merge_cells('A1:Z1')
sheet['A1'].alignment = Alignment(horizontal='left')
sheet['A1'].font = Font(bold=True)
start_row = 3

headers = list(average_prices.columns)
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=start_row, column=col_num)
    cell.value = header
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

for row in dataframe_to_rows(average_prices, index=False, header=False):
    sheet.append(row)
book.save(xlsx_file)
